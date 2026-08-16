# -*- coding: utf-8 -*-
"""
Corrige el encabezado de la hoja FEBRERO 2DA QUINCENA en los libros de GATMA.

QUÉ ESTABA MAL
  La hoja se llama 'FEBRERO 2DA QUINCENA' y sus facturas son del 17, 19, 23
  y 24 de febrero, pero su encabezado quedó copiado de la hoja anterior y
  decía '01/02/2026 al 15/02/2026'.

  No es cosmético. El período de declaración se lee del RANGO, no del nombre
  de la hoja, así que esas cuatro facturas —1.390.822,44 de base— entrarían a
  la primera quincena: la de febrero saldría inflada y la segunda vacía.

QUÉ TOCA Y QUÉ NO
  Solamente esa celda de texto. Ni un monto, ni una factura, ni una fórmula.
  El libro se abre con las fórmulas intactas (no con sus valores calculados),
  de modo que los SUM de los totales siguen siendo fórmulas después de
  guardar y no quedan congelados en el número que tenían hoy.

  Antes de escribir deja una copia con la fecha y hora en el nombre. Es un
  documento fiscal: si algo sale distinto a lo esperado, se vuelve atrás
  copiando el respaldo encima.

USO
  python corregir_febrero_gatma.py            (muestra qué haría)
  python corregir_febrero_gatma.py --aplicar
"""
import os
import re
import shutil
import sys
from datetime import datetime

CARPETA = r'C:\Users\torre\OneDrive\Escritorio\GATMA\libros auxiliares'
LIBROS = ['Libro de Ventas GATMA.xlsx', 'Libro de Compras GATMA.xlsx']
HOJA = 'FEBRERO 2DA QUINCENA'
# Febrero de 2026 tiene 28 días: 2026 no es bisiesto.
CORRECTO = '16/02/2026 al 28/02/2026'
RANGO = re.compile(r'\d{1,2}/\d{1,2}/\d{4}\s*al\s*\d{1,2}/\d{1,2}/\d{4}')


def main():
    aplicar = '--aplicar' in sys.argv
    try:
        import openpyxl
    except ImportError:
        sys.exit('Hace falta openpyxl:  pip install openpyxl')

    sello = datetime.now().strftime('%Y%m%d-%H%M%S')
    hubo = False

    for nombre in LIBROS:
        ruta = os.path.join(CARPETA, nombre)
        if not os.path.exists(ruta):
            print('  NO ESTÁ  %s' % nombre)
            continue

        # data_only en False a propósito: así se conservan las fórmulas. Con
        # True se leerían sus valores calculados y al guardar quedarían
        # escritos como números fijos, y los totales dejarían de sumar.
        libro = openpyxl.load_workbook(ruta)
        if HOJA not in libro.sheetnames:
            print('  sin la hoja %s: %s' % (HOJA, nombre))
            continue
        hoja = libro[HOJA]

        objetivo = None
        for fila in hoja.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str) and RANGO.search(celda.value):
                    objetivo = celda
                    break
            if objetivo:
                break

        if not objetivo:
            print('  sin encabezado de rango: %s' % nombre)
            continue

        antes = objetivo.value
        despues = RANGO.sub(CORRECTO, antes, count=1)
        print('\n%s · %s' % (nombre, objetivo.coordinate))
        print('   antes:   %s' % antes)
        print('   después: %s' % despues)

        if antes == despues:
            print('   (ya estaba correcto)')
            continue
        hubo = True

        if aplicar:
            # Se comprueba que el archivo se pueda escribir ANTES de copiarlo.
            # Excel lo mantiene bloqueado mientras está abierto, y sin esta
            # comprobación el respaldo se creaba, el guardado fallaba, y
            # quedaba una copia suelta de un archivo que nadie modificó.
            try:
                with open(ruta, 'r+b'):
                    pass
            except (PermissionError, OSError):
                print('   BLOQUEADO · ciérralo en Excel y vuelve a correrlo. '
                      'No se tocó nada.')
                continue

            respaldo = ruta.replace('.xlsx', ' (respaldo %s).xlsx' % sello)
            shutil.copy2(ruta, respaldo)
            objetivo.value = despues
            libro.save(ruta)
            print('   CORREGIDO · respaldo en: %s' % os.path.basename(respaldo))

    if not aplicar and hubo:
        print('\nEnsayo: no se escribió nada. Agrega --aplicar para hacerlo.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
